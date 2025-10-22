import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
import os
import msal
import requests
from typing import List, Dict, Optional
import glob
import sys

# Load environment variables
load_dotenv()

def find_csv_file() -> Optional[str]:
    """
    Find a CSV file that starts with 'Files_' in the current directory.
    Returns the file path if found, None otherwise.
    """
    csv_files = glob.glob('Files_*.csv')
    if csv_files:
        return csv_files[0]  # Return the first match
    return None

def get_csv_file_path() -> str:
    """
    Get the CSV file path either by auto-detection or user input.
    Validates that the file exists and is readable.
    """
    # Try to auto-detect
    csv_file = find_csv_file()

    if csv_file:
        print(f"  → Found CSV file: {csv_file}")
        return csv_file

    # No file found, prompt user
    print("\n  ⚠ No CSV file starting with 'Files_' found in current directory.")

    while True:
        csv_file = input("  → Please enter the path to your SharePoint permissions CSV file: ").strip()

        if not csv_file:
            print("  ✗ File path cannot be empty. Please try again.")
            continue

        # Remove quotes if user wrapped path in quotes
        csv_file = csv_file.strip('"').strip("'")

        if not os.path.exists(csv_file):
            print(f"  ✗ File not found: {csv_file}")
            retry = input("  → Try again? (y/n): ").strip().lower()
            if retry != 'y':
                print("\n✗ Cannot proceed without a valid CSV file. Exiting.")
                sys.exit(1)
            continue

        # Validate it's a CSV
        if not csv_file.lower().endswith('.csv'):
            print(f"  ✗ File must be a CSV file (has .csv extension)")
            retry = input("  → Try again? (y/n): ").strip().lower()
            if retry != 'y':
                print("\n✗ Cannot proceed without a valid CSV file. Exiting.")
                sys.exit(1)
            continue

        # Try to read the file to validate it
        try:
            test_df = pd.read_csv(csv_file, nrows=1)
            print(f"  ✓ Valid CSV file found: {csv_file}")
            return csv_file
        except Exception as e:
            print(f"  ✗ Error reading CSV file: {e}")
            retry = input("  → Try again? (y/n): ").strip().lower()
            if retry != 'y':
                print("\n✗ Cannot proceed without a valid CSV file. Exiting.")
                sys.exit(1)

def print_credential_help():
    """Print helpful information about obtaining Azure credentials."""
    print("\n" + "=" * 60)
    print("HOW TO OBTAIN Entra ID CREDENTIALS")
    print("=" * 60)
    print("1. Navigate to [Azure Portal - App Registrations](https://portal.azure.com/#blade/Microsoft_AAD_IAM/ActiveDirectoryMenuBlade/RegisteredApps)")
    print("2. Click '+ New registration'")
    print("3. After creating the app, note the:")
    print("   - Directory (tenant) ID")
    print("   - Application (client) ID")
    print("4. Go to 'Certificates & secrets' > '+ New client secret'")
    print("5. Copy the secret VALUE immediately (shown only once)")
    print("6. Go to 'API permissions' > '+ Add a permission'")
    print("7. Select 'Microsoft Graph' > 'Application permissions'")
    print("8. Add these permissions:")
    print("   - GroupMember.Read.All")
    print("   - User.Read.All")
    print("9. Click 'Grant admin consent'")
    print("=" * 60)

def validate_azure_credentials(tenant_id: str, client_id: str, client_secret: str) -> tuple:
    """
    Validate credential format before attempting authentication.
    Returns (is_valid, error_message)
    """
    # Check if fields are empty
    if not tenant_id or not tenant_id.strip():
        return False, "Tenant ID cannot be empty"
    if not client_id or not client_id.strip():
        return False, "Client ID cannot be empty"
    if not client_secret or not client_secret.strip():
        return False, "Client Secret cannot be empty"

    # Basic format validation for GUIDs (tenant and client IDs)
    guid_pattern = re.compile(r'^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$')

    if not guid_pattern.match(tenant_id.strip()):
        return False, "Tenant ID must be a valid GUID format (e.g., xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx)"

    if not guid_pattern.match(client_id.strip()):
        return False, "Client ID must be a valid GUID format (e.g., xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx)"

    return True, None

def prompt_manual_credentials() -> tuple:
    """
    Prompt user for manual credential input with validation and retry.
    Returns tuple of (tenant_id, client_id, client_secret, success)
    """
    print("\n  Please enter your Entra ID application credentials:")
    print("  (Type 'help' for instructions on obtaining credentials)")

    while True:
        tenant_id = input("\n  → Tenant ID (or 'help'/'cancel'): ").strip()

        if tenant_id.lower() == 'help':
            print_credential_help()
            continue
        elif tenant_id.lower() == 'cancel':
            return None, None, None, False

        client_id = input("  → Client ID: ").strip()
        client_secret = input("  → Client Secret: ").strip()

        # Validate credentials format
        is_valid, error_msg = validate_azure_credentials(tenant_id, client_id, client_secret)

        if not is_valid:
            print(f"\n  ✗ Invalid credentials: {error_msg}")
            retry = input("  → Try again? (y/n): ").strip().lower()
            if retry != 'y':
                return None, None, None, False
            continue

        return tenant_id, client_id, client_secret, True

def setup_graph_client():
    """
    Set up Graph API client with comprehensive error handling and retry logic.
    Returns an initialized GraphAPIClient (may be disabled if credentials not provided).
    """
    # First, try credentials from .env file
    tenant_id = os.getenv('AZURE_TENANT_ID')
    client_id = os.getenv('AZURE_CLIENT_ID')
    client_secret = os.getenv('AZURE_CLIENT_SECRET')

    credentials_source = "environment"
    has_env_creds = all([tenant_id, client_id, client_secret])

    if has_env_creds:
        print("  → Found credentials in .env file")

        # Validate format first
        is_valid, error_msg = validate_azure_credentials(tenant_id, client_id, client_secret)
        if not is_valid:
            print(f"  ✗ Invalid credentials in .env file: {error_msg}")
            print("  → Please check your .env file or provide credentials manually")
            has_env_creds = False
        else:
            # Try to authenticate
            print("  → Authenticating with Microsoft Graph API...")
            graph_client = GraphAPIClient(tenant_id, client_id, client_secret)

            if graph_client.enabled:
                print("  ✓ Successfully authenticated!")
                return graph_client
            else:
                # Authentication failed
                print(f"  ✗ Authentication failed: {graph_client.get_auth_error()}")

    # If env credentials failed or don't exist, offer manual input
    if not has_env_creds or (has_env_creds and not graph_client.enabled):
        while True:
            print("\n  What would you like to do?")
            print("  1. Enter credentials manually")
            print("  2. Continue without security group expansion")
            print("  3. Exit program")

            choice = input("\n  → Enter choice (1/2/3): ").strip()

            if choice == '1':
                # Manual credential input
                tenant_id, client_id, client_secret, success = prompt_manual_credentials()

                if not success:
                    print("  → Credential input cancelled")
                    continue

                # Try to authenticate with manual credentials
                print("\n  → Authenticating with provided credentials...")
                graph_client = GraphAPIClient(tenant_id, client_id, client_secret)

                if graph_client.enabled:
                    print("  ✓ Successfully authenticated!")
                    return graph_client
                else:
                    print(f"  ✗ Authentication failed: {graph_client.get_auth_error()}")
                    print("\n  Please check your credentials and try again.")

                    retry_choice = input("\n  → Try again? (y=retry, n=continue without, e=exit): ").strip().lower()

                    if retry_choice == 'y':
                        continue  # Loop back to credential input
                    elif retry_choice == 'n':
                        print("  → Proceeding without security group expansion")
                        return GraphAPIClient()  # Return disabled client
                    else:
                        print("\n✗ Exiting program")
                        sys.exit(0)

            elif choice == '2':
                print("  → Proceeding without security group expansion")
                return GraphAPIClient()  # Return disabled client

            elif choice == '3':
                print("\n✗ Exiting program")
                sys.exit(0)

            else:
                print("  ✗ Invalid choice. Please enter 1, 2, or 3.")

class GraphAPIClient:
    """
    Client for interacting with Microsoft Graph API to retrieve security group members.
    """
    def __init__(self, tenant_id=None, client_id=None, client_secret=None):
        self.tenant_id = tenant_id
        self.client_id = client_id
        self.client_secret = client_secret
        self.access_token = None
        self.group_cache = {}  # Cache to avoid repeated API calls for same groups
        self.auth_error = None  # Store authentication error details

        if not all([self.tenant_id, self.client_id, self.client_secret]):
            self.enabled = False
        else:
            self.enabled = True
            self._authenticate()

    def _authenticate(self):
        """
        Authenticate with Entra ID and get access token.
        Returns True on success, False on failure.
        """
        try:
            authority = f"https://login.microsoftonline.com/{self.tenant_id}"
            app = msal.ConfidentialClientApplication(
                self.client_id,
                authority=authority,
                client_credential=self.client_secret
            )

            result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])

            if "access_token" in result:
                self.access_token = result["access_token"]
                return True
            else:
                error_desc = result.get('error_description', 'Unknown error')
                error_code = result.get('error', 'Unknown')
                self.auth_error = f"{error_code}: {error_desc}"
                self.enabled = False
                return False
        except Exception as e:
            error_str = str(e)
            self.auth_error = error_str

            # Provide helpful error messages for common issues
            if "should consist of an https url" in error_str.lower():
                self.auth_error = "Invalid Tenant ID format. It should be a GUID (e.g., xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx)"
            elif "invalid_client" in error_str.lower():
                self.auth_error = "Invalid Client ID or Client Secret. Please verify your credentials."
            elif "unauthorized_client" in error_str.lower():
                self.auth_error = "Application not authorized. Ensure API permissions are granted and admin consent is provided."

            self.enabled = False
            return False

    def get_auth_error(self) -> str:
        """Get the authentication error message if authentication failed."""
        return self.auth_error if self.auth_error else "Unknown authentication error"

    def is_security_group(self, principal_name: str) -> bool:
        """
        Determine if a principal name represents a security group.
        Security groups typically don't contain '@' or end with specific patterns.
        """
        if not principal_name or pd.isna(principal_name):
            return False

        principal_str = str(principal_name).strip()

        # Security groups usually don't have @ symbol (unlike user emails)
        # and often have patterns like "SG_", "Group_", or are simple names
        if '@' not in principal_str:
            # Exclude system accounts
            if principal_str.lower() not in ['everyone', 'nt authority\\authenticated users',
                                             'system', 'sharepoint app', 'limited access system account']:
                return True

        return False

    def get_group_id_by_name(self, group_name: str) -> Optional[str]:
        """
        Get the group ID from Microsoft Graph API by display name.
        """
        if not self.enabled or not self.access_token:
            return None

        try:
            from urllib.parse import quote

            headers = {'Authorization': f'Bearer {self.access_token}'}
            # Properly URL encode the group name to handle special characters like &
            encoded_group_name = quote(group_name, safe='')
            url = f"https://graph.microsoft.com/v1.0/groups?$filter=displayName eq '{encoded_group_name}'"

            response = requests.get(url, headers=headers)
            response.raise_for_status()

            data = response.json()
            if data.get('value') and len(data['value']) > 0:
                return data['value'][0]['id']

            return None
        except Exception as e:
            print(f"Error fetching group ID for '{group_name}': {e}")
            return None

    def get_group_members(self, group_name: str) -> List[Dict[str, str]]:
        """
        Get all members of a security group from Microsoft Graph API.
        Returns a list of dictionaries with member information.
        """
        if not self.enabled:
            return []

        # Check cache first
        if group_name in self.group_cache:
            return self.group_cache[group_name]

        try:
            # Get group ID
            group_id = self.get_group_id_by_name(group_name)
            if not group_id:
                # Silently cache and return empty - group not found or is not an Entra ID group
                self.group_cache[group_name] = []
                return []

            headers = {'Authorization': f'Bearer {self.access_token}'}
            url = f"https://graph.microsoft.com/v1.0/groups/{group_id}/members"

            members = []
            while url:
                response = requests.get(url, headers=headers)
                response.raise_for_status()

                data = response.json()

                for member in data.get('value', []):
                    member_type = member.get('@odata.type', '').split('.')[-1]

                    # Skip SharePointGroup types
                    if member_type == 'group':
                        # Check if it's a SharePoint group by examining additional properties
                        # SharePoint groups typically don't have certain Entra ID properties
                        continue

                    member_info = {
                        'displayName': member.get('displayName', 'N/A'),
                        'userPrincipalName': member.get('userPrincipalName', 'N/A'),
                        'mail': member.get('mail', 'N/A'),
                        'memberType': member_type
                    }
                    members.append(member_info)

                # Handle pagination
                url = data.get('@odata.nextLink')

            self.group_cache[group_name] = members
            if members:
                print(f"  → Found security group: '{group_name}' ({len(members)} members)")
            return members

        except Exception as e:
            print(f"Error fetching members for group '{group_name}': {e}")
            self.group_cache[group_name] = []
            return []

def sanitize_sheet_name(name):
    """
    Removes illegal characters and truncates a string to make it a valid
    Excel sheet name.
    """
    # Remove the initial 'sites/Files/' prefix for brevity
    name = name.replace('sites/Files/', '')
    # Remove illegal characters: [ ] * ? : / \
    name = re.sub(r'[\[\]\*?:\/\\"]', '', name)
    # Truncate to Excel's 31-character limit
    return name[:31]

def create_excel_report_with_tables(data_groups, excel_file_name, graph_client=None):
    """
    Creates a styled Excel report from grouped pandas DataFrames, with each
    group on a separate sheet formatted as a table.

    Args:
        data_groups (pandas.core.groupby.DataFrameGroupBy): The grouped DataFrame.
        excel_file_name (str): The name of the Excel file to create.
        graph_client (GraphAPIClient, optional): Client for expanding security groups.
    """
    with pd.ExcelWriter(excel_file_name, engine='openpyxl') as writer:
        for group_name, group_df in data_groups:
            # Create a valid sheet name
            sheet_name = sanitize_sheet_name(group_name)

            # Write the dataframe to the sheet
            group_df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Access the worksheet object
            ws = writer.sheets[sheet_name]

            # --- Create and Apply Excel Table ---
            # Define the table range
            table_range = f"A1:{get_column_letter(group_df.shape[1])}{group_df.shape[0] + 1}"

            # Create a table object
            table = Table(displayName=f"Table_{sheet_name.replace(' ', '_')}", ref=table_range)

            # Add a style to the table
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)

            # --- Adjust Column Widths ---
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width

            # --- Expand Security Groups ---
            if graph_client and graph_client.enabled:
                expand_security_groups_in_sheet(ws, group_df, graph_client)

    print(f"Successfully created {excel_file_name}")

def expand_security_groups_in_sheet(ws, df, graph_client):
    """
    Identify security groups in the worksheet and add member information.

    Args:
        ws: Openpyxl worksheet object
        df: The DataFrame containing the data
        graph_client: GraphAPIClient instance
    """
    # Look for columns that might contain principal/user information
    principal_columns = []
    for idx, col_name in enumerate(df.columns, start=1):
        col_lower = str(col_name).lower()
        if any(keyword in col_lower for keyword in ['principal', 'user', 'name', 'account', 'granted']):
            principal_columns.append((idx, col_name))

    if not principal_columns:
        return

    # Styles for security group members section
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    member_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='top')

    # Find the rightmost column and the row after the table
    last_col = df.shape[1]
    start_row = df.shape[0] + 3  # Leave some space after the table

    security_groups_found = {}

    # Scan the data for security groups
    for row_idx in range(2, df.shape[0] + 2):  # Starting from row 2 (after headers)
        for col_idx, col_name in principal_columns:
            cell_value = ws.cell(row=row_idx, column=col_idx).value

            if graph_client.is_security_group(cell_value):
                group_name = str(cell_value).strip()

                if group_name not in security_groups_found:
                    members = graph_client.get_group_members(group_name)
                    if members:
                        security_groups_found[group_name] = members

    # Add security group information to the sheet
    if security_groups_found:
        current_row = start_row

        # Add section title
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.value = "Security Group Members"
        title_cell.font = Font(size=14, bold=True, color="366092")
        current_row += 2

        for group_name, members in security_groups_found.items():
            # Group header
            group_header_cell = ws.cell(row=current_row, column=1)
            group_header_cell.value = f"Group: {group_name}"
            group_header_cell.font = Font(size=12, bold=True)
            group_header_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            current_row += 1

            # Member table headers
            headers = ["Display Name", "Email", "Member Type"]
            for col_offset, header in enumerate(headers, start=1):
                header_cell = ws.cell(row=current_row, column=col_offset)
                header_cell.value = header
                header_cell.fill = header_fill
                header_cell.font = header_font
                header_cell.alignment = center_alignment

            current_row += 1

            # Member data
            for member in members:
                ws.cell(row=current_row, column=1).value = member.get('displayName', 'N/A')
                ws.cell(row=current_row, column=2).value = member.get('mail', 'N/A')
                ws.cell(row=current_row, column=3).value = member.get('memberType', 'N/A')

                # Apply styling to member rows
                for col in range(1, 4):
                    cell = ws.cell(row=current_row, column=col)
                    cell.fill = member_fill
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

                current_row += 1

            # Add spacing between groups
            current_row += 2

        # Adjust column widths for security group section
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20

        # Summary message already printed by individual group discoveries
        pass

# --- Main Script ---
if __name__ == "__main__":
    try:
        print("=" * 60)
        print("SharePoint Permissions Report Generator")
        print("=" * 60)

        # Get CSV file path
        print("\n[1/4] Locating CSV file...")
        csv_file_path = get_csv_file_path()

        # Setup Azure credentials and Graph API client
        print("\n[2/4] Setting up Azure credentials...")
        graph_client = setup_graph_client()

        print(f"\n[3/4] Reading CSV file...")
        df = pd.read_csv(csv_file_path)
        print(f"  → Loaded {len(df)} permission entries")

        # --- 1. Generate Parent Permissions Report ---
        print("\n[4/5] Generating Parent Permissions Report...")
        parent_df = df[df['Item Type'].isin(['Web', 'List'])].copy()
        # Drop columns that are often empty for parent libraries
        parent_df = parent_df.drop(columns=['Link ID', 'Link Type', 'AccessViaLinkID'], errors='ignore')
        parent_groups = parent_df.groupby('Resource Path')
        print(f"  → Processing {len(parent_groups)} parent resources")
        create_excel_report_with_tables(parent_groups, 'Parent_Permissions.xlsx', graph_client)

        # --- 2. Generate Unique Permissions Report ---
        print("\n[5/5] Generating Unique Permissions Report...")
        unique_df = df[~df['Item Type'].isin(['Web', 'List'])].copy()

        # Identify the parent library for each unique permission entry
        parent_paths = parent_df['Resource Path'].unique()
        def find_parent_library(path):
            for parent in parent_paths:
                if path.startswith(parent):
                    return parent
            # Fallback if no parent is found (e.g., top-level unique items)
            return '/'.join(path.split('/')[:3])

        if not unique_df.empty:
            unique_df['Document Library'] = unique_df['Resource Path'].apply(find_parent_library)
            unique_groups = unique_df.groupby('Document Library')
            print(f"  → Processing {len(unique_groups)} document libraries with unique permissions")
            create_excel_report_with_tables(unique_groups, 'Unique_Permissions.xlsx', graph_client)
        else:
            print("  → No unique permissions found")

        print("\n" + "=" * 60)
        print("Report generation complete!")
        print("=" * 60)

    except FileNotFoundError:
        print(f"Error: The file '{csv_file_path}' was not found. Please ensure it's in the same directory.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        import traceback
        traceback.print_exc()